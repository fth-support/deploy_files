import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import os
import subprocess
import threading
from datetime import datetime

PSEXEC_PATH = r"C:\Support\PSTools\PsExec.exe"

# ค่าคงที่สำหรับซ่อนหน้าต่าง CMD บน Windows
CREATE_NO_WINDOW = 0x08000000

class DeployApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hybrid Mass Deployment Tool")
        self.root.geometry("1100x750") # ขยายหน้าต่างขึ้นนิดหน่อยเพื่อรองรับคอลัมน์ใหม่
        
        self.source_path = tk.StringVar()
        self.exe_path = tk.StringVar()
        
        self.setup_ui()

    def setup_ui(self):
        # ================= ส่วนที่ 1.1: Browse Files =================
        frame1 = tk.LabelFrame(self.root, text="1. Source & Execute Settings", padx=10, pady=10)
        frame1.pack(fill="x", padx=10, pady=5)

        tk.Label(frame1, text="Source (File/Folder):").grid(row=0, column=0, sticky="w")
        tk.Entry(frame1, textvariable=self.source_path, width=70).grid(row=0, column=1, padx=5)
        tk.Button(frame1, text="Browse Folder", command=lambda: self.source_path.set(filedialog.askdirectory())).grid(row=0, column=2, padx=2)
        tk.Button(frame1, text="Browse File", command=lambda: self.source_path.set(filedialog.askopenfilename())).grid(row=0, column=3, padx=2)

        tk.Label(frame1, text="EXE to Run (Optional):").grid(row=1, column=0, sticky="w", pady=5)
        tk.Entry(frame1, textvariable=self.exe_path, width=70).grid(row=1, column=1, padx=5)
        tk.Button(frame1, text="Browse EXE", command=lambda: self.exe_path.set(filedialog.askopenfilename(filetypes=[("Executable", "*.exe")]))).grid(row=1, column=2, columnspan=2, sticky="ew", padx=2)

        # ================= ส่วนที่ 1.2: Excel Template =================
        frame2 = tk.LabelFrame(self.root, text="2. Import / Export Data", padx=10, pady=10)
        frame2.pack(fill="x", padx=10, pady=5)

        tk.Button(frame2, text="Export Template (Excel)", command=self.export_template).pack(side="left", padx=5)
        tk.Button(frame2, text="Import Data (Excel)", command=self.import_excel).pack(side="left", padx=5)
        tk.Button(frame2, text="Start Deployment", bg="green", fg="white", font=("Arial", 10, "bold"), command=self.start_deployment).pack(side="right", padx=5)

        # ================= ส่วนที่ 1.3: Notebook Tabs (Targets & Logs) =================
        notebook_frame = tk.Frame(self.root)
        notebook_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.notebook = ttk.Notebook(notebook_frame)
        self.notebook.pack(fill="both", expand=True)

        # --- Tab 1: ตาราง Targets ---
        tab_targets = ttk.Frame(self.notebook)
        self.notebook.add(tab_targets, text="Deployment Targets & Status")

        # เพิ่มคอลัมน์ Progress
        columns = ("IP", "Username", "Password", "Run_EXE", "Custom_Source", "Progress", "Ping", "Copy", "Run")
        self.tree = ttk.Treeview(tab_targets, columns=columns, show="headings", selectmode="browse")
        
        for col in columns:
            self.tree.heading(col, text=col)
            width = 120 if col == "Progress" else 80
            self.tree.column(col, width=width, anchor="center")
        self.tree.column("Custom_Source", width=150)
        self.tree.pack(fill="both", expand=True, pady=5)

        edit_frame = tk.Frame(tab_targets)
        edit_frame.pack(fill="x", pady=5)
        tk.Button(edit_frame, text="Edit Selected Row (Override)", command=self.edit_selected_row).pack(side="left")
        tk.Button(edit_frame, text="Clear Selected", command=self.delete_selected_row).pack(side="left", padx=5)

        # --- Tab 2: หน้าต่าง Logs ---
        tab_logs = ttk.Frame(self.notebook)
        self.notebook.add(tab_logs, text="📝 Execution Logs")
        
        log_scroll = tk.Scrollbar(tab_logs)
        log_scroll.pack(side="right", fill="y")
        
        self.log_text = tk.Text(tab_logs, wrap="word", yscrollcommand=log_scroll.set, state="disabled", bg="black", fg="lightgreen", font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True)
        log_scroll.config(command=self.log_text.yview)

        # ================= ส่วนที่ 1.5: Summary =================
        self.summary_var = tk.StringVar()
        self.summary_var.set("Total: 0 | 🟢 Success: 0 | 🔴 Fail: 0")
        tk.Label(self.root, textvariable=self.summary_var, font=("Arial", 12, "bold")).pack(pady=10)

    # --- ฟังก์ชันเขียน Log ---
    def write_log(self, text, level="INFO"):
        time_str = datetime.now().strftime("%H:%M:%S")
        log_msg = f"[{time_str}] [{level}] {text}\n"
        
        def append():
            self.log_text.config(state="normal")
            self.log_text.insert("end", log_msg)
            # ไฮไลท์สีถ้าเป็น ERROR
            if level == "ERROR":
                pos_start = self.log_text.index("end-2c linestart")
                pos_end = self.log_text.index("end-1c")
                self.log_text.tag_add("error", pos_start, pos_end)
                self.log_text.tag_config("error", foreground="red")
            
            self.log_text.see("end")
            self.log_text.config(state="disabled")
            
        self.root.after(0, append)

    # --- ฟังก์ชันจัดการ Excel ---
    def export_template(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="DeployTemplate.xlsx")
        if filepath:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["IP", "Username", "Password", "Run_EXE (Yes/No)", "Custom_Source_Path (Optional)"])
            ws.append(["192.168.1.50", "Administrator", "P@ssw0rd", "Yes", ""])
            wb.save(filepath)
            messagebox.showinfo("Success", "Template exported successfully!")

    def import_excel(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            try:
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and str(row[0]).strip() != "":
                        ip, user, pw = row[0], row[1] or "", row[2] or ""
                        run_opt = row[3] or "No"
                        custom_src = row[4] or ""
                        
                        # แทรกคอลัมน์ Progress (0%) เพิ่มเข้ามา
                        self.tree.insert("", "end", values=(ip, user, pw, run_opt, custom_src, "[          ] 0%", "⚪", "⚪", "⚪"))
                
                self.update_summary()
                self.write_log(f"Imported configuration from Excel: {os.path.basename(filepath)}")
                
            except Exception as e:
                messagebox.showerror("Import Error", f"เกิดข้อผิดพลาดในการโหลดไฟล์ Excel:\n{str(e)}")

    # --- ฟังก์ชันอัปเดต UI (อัปเดตสถานะและหลอด Progress) ---
    def set_progress(self, item_id, percent):
        bars = int(percent / 10)
        bar_str = "[" + "■" * bars + " " * (10 - bars) + f"] {percent}%"
        self.update_status(item_id, 5, bar_str) # Index 5 คือคอลัมน์ Progress

    def update_status(self, item_id, col_index, val):
        vals = list(self.tree.item(item_id)['values'])
        vals[col_index] = val
        self.tree.item(item_id, values=vals)
        self.root.update()

    def edit_selected_row(self):
        # (ฟังก์ชันเหมือนเดิม แต่แก้ Index ให้ตรงกับคอลัมน์ที่เพิ่มใหม่)
        selected = self.tree.selection()
        if not selected: return
        item = self.tree.item(selected[0])['values']
        top = tk.Toplevel(self.root)
        top.title("Edit Override")
        labels = ["IP", "Username", "Password", "Run_EXE (Yes/No)", "Custom_Source"]
        entries = []
        for i, text in enumerate(labels):
            tk.Label(top, text=text).grid(row=i, column=0, padx=5, pady=5)
            e = tk.Entry(top, width=40)
            e.insert(0, str(item[i]) if item[i] else "")
            e.grid(row=i, column=1, padx=5, pady=5)
            entries.append(e)
            
        def save_edit():
            new_vals = [e.get() for e in entries] + item[5:]
            self.tree.item(selected[0], values=new_vals)
            top.destroy()
            self.update_summary()
            
        tk.Button(top, text="Save", command=save_edit).grid(row=5, column=1, sticky="e", pady=10)

    def delete_selected_row(self):
        selected = self.tree.selection()
        if selected:
            self.tree.delete(selected[0])
            self.update_summary()

    def update_summary(self):
        total = len(self.tree.get_children())
        success, fail = 0, 0
        for child in self.tree.get_children():
            vals = self.tree.item(child)['values']
            if "🔴" in vals: fail += 1
            elif "🟢" in vals[7]: success += 1 # เช็คที่ Copy Success
        self.summary_var.set(f"Total: {total} | 🟢 Success: {success} | 🔴 Fail: {fail}")

    # --- ฟังก์ชันรันคำสั่งแบบซ่อนหน้าจอและเก็บ Log ---
    def run_cmd_hidden(self, cmd):
        try:
            # subprocess.run ร่วมกับ CREATE_NO_WINDOW จะซ่อนหน้าจอ cmd ได้ 100%
            res = subprocess.run(cmd, shell=True, capture_output=True, text=True, creationflags=CREATE_NO_WINDOW)
            return res.returncode, res.stdout.strip(), res.stderr.strip()
        except Exception as e:
            return -1, "", str(e)

    def start_deployment(self):
        if len(self.tree.get_children()) == 0:
            messagebox.showwarning("Warning", "กรุณา Import ข้อมูล Excel ก่อนเริ่ม Deployment")
            return
            
        # เปิดหน้าต่าง Log อัตโนมัติเวลากด Start
        self.notebook.select(1) 
        self.write_log("=== STARTED DEPLOYMENT TASK ===", "INFO")
        threading.Thread(target=self._run_deployment_thread, daemon=True).start()

    def _run_deployment_thread(self):
        items = self.tree.get_children()
        for item_id in items:
            vals = self.tree.item(item_id)['values']
            ip, user, pw, run_opt, custom_src = vals[0], vals[1], vals[2], vals[3], vals[4]
            
            self.write_log(f"--- Processing IP: {ip} ---", "INFO")
            self.set_progress(item_id, 10)

            # 1. Ping Check
            self.update_status(item_id, 6, "🟡") 
            code, out, err = self.run_cmd_hidden(f"ping -n 1 -w 1000 {ip}")
            if code != 0:
                self.write_log(f"[{ip}] Ping failed. Host is unreachable.", "ERROR")
                self.update_status(item_id, 6, "🔴")
                self.update_status(item_id, 7, "🔴")
                self.update_status(item_id, 8, "🔴")
                self.set_progress(item_id, 10) # ค้างที่ 10%
                continue
            
            self.write_log(f"[{ip}] Ping success.", "INFO")
            self.update_status(item_id, 6, "🟢")
            self.set_progress(item_id, 30)

            # 2. Copy Files 
            self.update_status(item_id, 7, "🟡")
            src = custom_src if custom_src else self.source_path.get()
            
            if not src:
                self.write_log(f"[{ip}] Source path is empty. Skipping copy.", "WARN")
                self.update_status(item_id, 7, "⚪") 
            else:
                self.write_log(f"[{ip}] Authenticating and mapping drive...", "INFO")
                # Map drive
                code, out, err = self.run_cmd_hidden(f'net use \\\\{ip}\\IPC$ /user:{user} {pw}')
                if code != 0:
                    self.write_log(f"[{ip}] Authentication failed! Error: {err or out}", "ERROR")
                    self.update_status(item_id, 7, "🔴")
                    continue
                
                self.set_progress(item_id, 50)
                dest = f"\\\\{ip}\\C$\\TempDeploy"
                
                self.write_log(f"[{ip}] Copying files to {dest}...", "INFO")
                if os.path.isdir(src):
                    cmd_copy = f'robocopy "{src}" "{dest}" /E /njh /njs /nc /ns /np'
                else:
                    cmd_copy = f'echo F | xcopy /Y /F "{src}" "{dest}\\{os.path.basename(src)}"'
                
                code_copy, out_copy, err_copy = self.run_cmd_hidden(cmd_copy)
                
                # Unmap drive
                self.run_cmd_hidden(f'net use \\\\{ip}\\IPC$ /delete /y')

                if code_copy in [0, 1, 2, 3]: # Robocopy success codes
                    self.write_log(f"[{ip}] Copy success.", "INFO")
                    self.update_status(item_id, 7, "🟢")
                    self.set_progress(item_id, 70)
                else:
                    self.write_log(f"[{ip}] Copy failed. Code: {code_copy} | Output: {out_copy}\n{err_copy}", "ERROR")
                    self.update_status(item_id, 7, "🔴")
                    continue

            # 3. Run PsExec
            if str(run_opt).strip().lower() == 'yes':
                self.update_status(item_id, 8, "🟡")
                exe_target = self.exe_path.get()
                if exe_target:
                    exe_name = os.path.basename(exe_target)
                    remote_exe_path = f"C:\\TempDeploy\\{exe_name}"
                    
                    self.write_log(f"[{ip}] Executing {exe_name} via PsExec...", "INFO")
                    psexec_cmd = f'"{PSEXEC_PATH}" \\\\{ip} -u {user} -p {pw} -accepteula -s -d "{remote_exe_path}"'
                    
                    code_px, out_px, err_px = self.run_cmd_hidden(psexec_cmd)
                    
                    if code_px == 0:
                        self.write_log(f"[{ip}] Execute trigger success.", "INFO")
                        self.update_status(item_id, 8, "🟢")
                        self.set_progress(item_id, 100)
                    else:
                        # PsExec มักจะพ่น Error ออกทาง STDERR
                        self.write_log(f"[{ip}] PsExec failed. Output: {err_px or out_px}", "ERROR")
                        self.update_status(item_id, 8, "🔴")
                else:
                    self.write_log(f"[{ip}] No EXE selected to run.", "WARN")
                    self.update_status(item_id, 8, "🔴") 
            else:
                self.write_log(f"[{ip}] Run optional set to NO. Completed.", "INFO")
                self.update_status(item_id, 8, "⚪")
                self.set_progress(item_id, 100)

        self.update_summary()
        self.write_log("=== DEPLOYMENT TASK COMPLETED ===", "INFO")
        
        # สลับหน้าจอกลับมาที่ตาราง Targets 
        self.notebook.select(0)
        messagebox.showinfo("Done", "Deployment Task Completed! Check Logs for details.")

if __name__ == "__main__":
    root = tk.Tk()
    app = DeployApp(root)
    root.mainloop()
